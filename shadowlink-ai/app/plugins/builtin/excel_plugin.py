"""Excel processing plugin — extracts data from .xlsx files."""

from __future__ import annotations

from typing import Any

from app.models.mcp import PluginInfo, ToolCategory, ToolInfo
from app.plugins.interface import PluginInterface


class ExcelPlugin(PluginInterface):
    """Excel document processing plugin."""

    def get_info(self) -> PluginInfo:
        return PluginInfo(
            name="excel_processor",
            version="0.1.0",
            description="Extract data from Excel spreadsheets using openpyxl",
            author="ShadowLink",
            supported_formats=[".xlsx", ".xls"],
            tools_provided=["parse_excel"],
        )

    async def initialize(self, config: dict[str, Any] | None = None) -> None:
        pass

    async def shutdown(self) -> None:
        pass

    def get_tools(self) -> list[tuple[ToolInfo, Any]]:
        tool_info = ToolInfo(
            name="parse_excel",
            description="Parse an Excel file and extract sheet data as text",
            category=ToolCategory.FILE,
            input_schema={
                "type": "object",
                "properties": {
                    "file_path": {"type": "string", "description": "Path to Excel file"},
                    "sheet_name": {"type": "string", "description": "Sheet name (default: active sheet)"},
                },
                "required": ["file_path"],
            },
        )
        return [(tool_info, self._parse_excel)]

    @staticmethod
    async def _parse_excel(file_path: str, sheet_name: str | None = None) -> str:
        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append("\t".join(str(cell) if cell is not None else "" for cell in row))
            wb.close()
            return "\n".join(rows)
        except ImportError:
            return "Error: openpyxl not installed"
        except Exception as exc:
            return f"Error processing Excel: {exc}"


def create_plugin() -> PluginInterface:
    return ExcelPlugin()
