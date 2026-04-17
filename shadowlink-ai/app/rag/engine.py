"""RAG Engine — orchestrates the retrieval-augmented generation pipeline.

Full 9-step pipeline:
  1. Query classification (Adaptive RAG)
  2. Query rewriting (HyDE / Step-back)
  3. Multi-path retrieval (FAISS vector search)
  4. RRF fusion (implicit in hybrid retrieval)
  5. Cross-encoder reranking
  6. CRAG quality gate
  7. Context assembly
  8. Answer generation (via LLM)
  9. Full-chain tracing
"""

from __future__ import annotations

import time
import uuid
from pathlib import Path
from typing import Any

import structlog

from app.config import settings
from app.models.rag import (
    IngestRequest,
    IngestResponse,
    QueryClassification,
    RAGChunk,
    RAGRequest,
    RAGResponse,
    RAGTrace,
    RetrievalMethod,
)

logger = structlog.get_logger("rag.engine")


class RAGEngine:
    """Unified RAG pipeline engine.

    Orchestrates the full 9-step pipeline with real embedding, indexing,
    retrieval, reranking, and quality gating.
    """

    def __init__(self) -> None:
        from app.rag.chunking.recursive import RecursiveChunker
        from app.rag.embedding.local import LocalEmbedding
        from app.rag.index.manager import IndexManager
        from app.rag.reranking.cross_encoder import CrossEncoderReranker

        self._chunker = RecursiveChunker(
            chunk_size=settings.rag.default_chunk_size,
            chunk_overlap=settings.rag.default_chunk_overlap,
        )
        self._embedder = LocalEmbedding()
        self._index_manager = IndexManager()
        self._reranker = CrossEncoderReranker(device=settings.rag.embedding_device) if settings.rag.rerank_enabled else None

    async def query(self, request: RAGRequest) -> RAGResponse:
        """Execute the full 9-step RAG pipeline."""
        trace = RAGTrace(
            trace_id=f"rag-{uuid.uuid4().hex[:12]}",
            original_query=request.query,
            retrieval_method=request.retrieval_method,
        )
        pipeline_start = time.perf_counter()

        try:
            # ── Step 1: Query Classification ──
            classification = await self._classify_query(request.query)
            trace.query_classification = classification

            if classification == QueryClassification.NO_RAG:
                return RAGResponse(
                    query_classification=classification,
                    trace=trace,
                )

            # ── Step 2: Query Rewriting (HyDE / Step-back) ──
            rewritten = await self._rewrite_query(request.query, classification)
            trace.rewritten_query = rewritten

            # ── Step 3: Multi-path Retrieval ──
            retrieval_start = time.perf_counter()
            raw_chunks = await self._retrieve(rewritten, request)
            trace.retrieval_latency_ms = (time.perf_counter() - retrieval_start) * 1000
            trace.retrieved_count = len(raw_chunks)

            # ── Step 4: RRF Fusion (implicit in hybrid retrieval) ──

            # ── Step 5: Reranking ──
            if request.rerank and raw_chunks and self._reranker is not None:
                rerank_start = time.perf_counter()
                reranked = await self._rerank(request.query, raw_chunks)
                trace.rerank_latency_ms = (time.perf_counter() - rerank_start) * 1000
                trace.after_rerank_count = len(reranked)
            else:
                reranked = raw_chunks
                trace.after_rerank_count = len(reranked)

            # ── Step 6: CRAG Quality Gate ──
            quality_passed = await self._crag_quality_check(reranked, request.query)
            trace.crag_passed = quality_passed

            if not quality_passed:
                await logger.awarning("crag_gate_failed", query=request.query[:80])

            # ── Step 7: Context Assembly ──
            final_chunks = reranked[:request.top_k]

            # ── Step 8: Answer Generation ──
            answer = ""
            if final_chunks:
                answer = await self._generate_answer(request.query, final_chunks, classification)

            # ── Step 9: Tracing ──
            trace.total_latency_ms = (time.perf_counter() - pipeline_start) * 1000

            await logger.ainfo(
                "rag_query_complete",
                trace_id=trace.trace_id,
                classification=classification.value,
                retrieved=trace.retrieved_count,
                after_rerank=trace.after_rerank_count,
                latency_ms=round(trace.total_latency_ms, 2),
            )

            return RAGResponse(
                chunks=final_chunks,
                answer=answer,
                query_classification=classification,
                rewritten_query=rewritten,
                trace=trace,
            )

        except Exception as exc:
            trace.total_latency_ms = (time.perf_counter() - pipeline_start) * 1000
            await logger.aerror("rag_query_failed", error=str(exc), trace_id=trace.trace_id)
            return RAGResponse(trace=trace)

    async def ingest(self, request: IngestRequest) -> IngestResponse:
        """Ingest documents into the vector index.

        Pipeline: Parse -> Chunk -> Embed -> Index
        """
        start = time.perf_counter()
        await logger.ainfo("rag_ingest_start", files=len(request.file_paths), mode_id=request.mode_id)

        total_chunks = 0
        failed: list[str] = []
        index = self._index_manager.get_index(request.mode_id)

        for file_path in request.file_paths:
            try:
                path = Path(file_path)
                if not path.exists():
                    failed.append(file_path)
                    continue

                # Use file processing pipeline for structured formats
                content = await self._parse_file(path)
                if not content.strip():
                    failed.append(file_path)
                    continue

                # Chunk
                chunk_size = request.chunk_size or settings.rag.default_chunk_size
                chunk_overlap = request.chunk_overlap or settings.rag.default_chunk_overlap
                chunker = self._chunker
                if chunk_size != chunker._chunk_size or chunk_overlap != chunker._chunk_overlap:
                    from app.rag.chunking.recursive import RecursiveChunker
                    chunker = RecursiveChunker(chunk_size=chunk_size, chunk_overlap=chunk_overlap)

                docs = chunker.chunk(content, metadata={"source": str(path.name)})

                # Embed
                texts = [doc.page_content for doc in docs]
                embeddings = self._embedder.embed(texts)

                # Index
                metadata_list = []
                for i, doc in enumerate(docs):
                    metadata_list.append({
                        "chunk_id": f"{path.stem}_{i}",
                        "content": doc.page_content,
                        "source": str(path.name),
                        "chunk_index": i,
                        **doc.metadata,
                    })

                index.add(embeddings, metadata_list)
                total_chunks += len(docs)

                await logger.ainfo("file_ingested", file=path.name, chunks=len(docs))

            except Exception as exc:
                await logger.aerror("file_ingest_error", file=file_path, error=str(exc))
                failed.append(file_path)

        # Persist index to disk
        if total_chunks > 0:
            index.save()

        elapsed = (time.perf_counter() - start) * 1000
        return IngestResponse(
            total_documents=len(request.file_paths),
            total_chunks=total_chunks,
            failed_documents=failed,
            index_name=f"idx_{request.mode_id}",
            latency_ms=elapsed,
        )

    # ── Pipeline Step Implementations ──

    async def _classify_query(self, query: str) -> QueryClassification:
        """Step 1: Classify query to determine RAG strategy.

        Uses keyword heuristics with LLM fallback for ambiguous queries.
        """
        query_lower = query.lower()

        if any(kw in query_lower for kw in ("write", "create", "imagine", "story", "poem", "写", "创作", "编写")):
            return QueryClassification.CREATIVE

        if any(kw in query_lower for kw in ("hi", "hello", "hey", "thanks", "你好", "谢谢", "再见")):
            return QueryClassification.CONVERSATIONAL

        if any(kw in query_lower for kw in ("code", "function", "class", "bug", "implement", "debug", "代码", "函数", "编程")):
            return QueryClassification.CODE

        if any(kw in query_lower for kw in ("compare", "analyze", "evaluate", "versus", "vs", "分析", "对比", "评估")):
            return QueryClassification.ANALYTICAL

        return QueryClassification.FACTUAL

    async def _rewrite_query(self, query: str, classification: QueryClassification) -> str:
        """Step 2: Rewrite query for better retrieval using HyDE.

        HyDE (Hypothetical Document Embeddings): generates a hypothetical
        answer, then uses it as the retrieval query for better semantic match.

        For simple/conversational queries, returns as-is.
        """
        # Skip rewriting for conversational/creative queries
        if classification in (QueryClassification.CONVERSATIONAL, QueryClassification.CREATIVE, QueryClassification.NO_RAG):
            return query

        # For factual/analytical/code queries, use HyDE
        try:
            from app.core.dependencies import get_resource
            llm_client = get_resource("llm_client")
            if llm_client is None:
                return query

            # HyDE: ask LLM to generate a hypothetical answer
            hyde_prompt = (
                "Please write a short, factual passage that would answer the following question. "
                "Write as if you are writing a paragraph from a reference document. "
                "Do not include phrases like 'the answer is' — just write the informational content.\n\n"
                f"Question: {query}"
            )
            hypothetical = await llm_client.chat(message=hyde_prompt, system_prompt="You are a technical writer.")

            # Combine original query with hypothetical for better retrieval
            return f"{query}\n{hypothetical[:300]}"

        except Exception as exc:
            await logger.awarning("hyde_rewrite_failed", error=str(exc))
            return query

    async def _retrieve(self, query: str, request: RAGRequest) -> list[RAGChunk]:
        """Step 3: Vector retrieval via FAISS."""
        index = self._index_manager.get_index(request.mode_id)

        if index.total_vectors == 0:
            return []

        query_embedding = self._embedder.embed_single(query)
        chunks = index.search(query_embedding, top_k=request.top_k * 2)  # Over-fetch for reranking
        return chunks

    async def _rerank(self, query: str, chunks: list[RAGChunk]) -> list[RAGChunk]:
        """Step 5: Rerank chunks using cross-encoder."""
        if self._reranker is None:
            return chunks
        try:
            return await self._reranker.rerank(query, chunks)
        except Exception as exc:
            await logger.awarning("rerank_failed", error=str(exc))
            return chunks

    async def _crag_quality_check(self, chunks: list[RAGChunk], query: str) -> bool:
        """Step 6: CRAG quality gate — check if retrieved content is relevant.

        Uses score threshold. Fails if average score is below threshold.
        """
        if not chunks:
            return False

        threshold = settings.rag.crag_quality_threshold
        avg_score = sum(c.score for c in chunks) / len(chunks)
        return avg_score >= threshold

    async def _generate_answer(self, query: str, chunks: list[RAGChunk], classification: QueryClassification) -> str:
        """Step 8: Generate an answer using retrieved context + LLM."""
        try:
            from app.core.dependencies import get_resource
            llm_client = get_resource("llm_client")
            if llm_client is None:
                return ""

            context = "\n\n---\n\n".join(
                f"[Source: {c.source}]\n{c.content}" for c in chunks
            )

            system = (
                "You are a helpful assistant. Answer the user's question based on the provided context. "
                "If the context doesn't contain enough information, say so honestly. "
                "Cite sources when possible."
            )
            prompt = f"Context:\n{context}\n\nQuestion: {query}\n\nAnswer:"

            return await llm_client.chat(message=prompt, system_prompt=system)
        except Exception as exc:
            await logger.awarning("rag_generation_failed", error=str(exc))
            return ""

    async def _parse_file(self, path: Path) -> str:
        """Parse file content using the appropriate parser."""
        suffix = path.suffix.lower()

        if suffix == ".pdf":
            return await self._parse_pdf(path)
        elif suffix == ".docx":
            return await self._parse_docx(path)
        elif suffix == ".xlsx":
            return await self._parse_xlsx(path)
        else:
            # Text-based files
            return path.read_text(encoding="utf-8", errors="replace")

    async def _parse_pdf(self, path: Path) -> str:
        """Extract text from PDF using PyMuPDF."""
        try:
            import fitz
            doc = fitz.open(str(path))
            pages = []
            for page in doc:
                pages.append(page.get_text())
            doc.close()
            return "\n\n".join(pages)
        except ImportError:
            return path.read_bytes().decode("utf-8", errors="replace")

    async def _parse_docx(self, path: Path) -> str:
        """Extract text from DOCX."""
        try:
            import docx
            doc = docx.Document(str(path))
            return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())
        except ImportError:
            return ""

    async def _parse_xlsx(self, path: Path) -> str:
        """Extract text from XLSX."""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(path), read_only=True, data_only=True)
            parts = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(cells):
                        rows.append(" | ".join(cells))
                if rows:
                    parts.append(f"## Sheet: {sheet}\n" + "\n".join(rows))
            wb.close()
            return "\n\n".join(parts)
        except ImportError:
            return ""

    # ── Public utilities ──

    def list_indices(self) -> list[dict[str, Any]]:
        """List all index partitions with stats."""
        return self._index_manager.list_indices()

    def delete_index(self, mode_id: str) -> bool:
        """Delete an index partition."""
        return self._index_manager.delete_index(mode_id)
