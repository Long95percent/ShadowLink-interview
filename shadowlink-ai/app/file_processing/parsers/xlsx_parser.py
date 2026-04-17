"""Excel parser using openpyxl."""

from __future__ import annotations

from pathlib import Path

from app.models.mcp import ExtractedTable, ParsedDocument


async def parse_xlsx(file_path: str | Path) -> ParsedDocument:
    """Parse an XLSX file and extract sheet data."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        text_parts = []
        tables = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                row_data = [str(cell) if cell is not None else "" for cell in row]
                rows.append(row_data)

            if rows:
                headers = rows[0] if rows else []
                text_parts.append(f"## Sheet: {sheet_name}\n")
                text_parts.append("\t".join(headers))
                for row in rows[1:]:
                    text_parts.append("\t".join(row))

                tables.append(ExtractedTable(
                    headers=headers,
                    rows=rows[1:],
                    caption=f"Sheet: {sheet_name}",
                ))

        wb.close()
        return ParsedDocument(
            source=str(file_path),
            content="\n".join(text_parts),
            file_type=".xlsx",
            tables=tables,
            metadata={"sheet_count": len(wb.sheetnames)},
        )
    except ImportError:
        return ParsedDocument(source=str(file_path), content="Error: openpyxl not installed", file_type=".xlsx")
